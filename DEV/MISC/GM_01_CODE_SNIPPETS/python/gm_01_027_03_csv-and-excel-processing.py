# PATTERN: CSV and Excel Processing

import csv
import os

# Setup: Create a dummy CSV file
with open("products.csv", "w", newline="") as f:
    f.write("ID,Name,Price\n")
    f.write("1,Laptop,1200\n")
    f.write("2,Mouse,25\n")

# Pattern: Read data using csv.DictReader
product_list = []
with open("products.csv", "r", newline="") as csvfile:
    reader = csv.DictReader(csvfile)
    for row in reader:
        product_list.append(row)

os.remove("products.csv") # Clean up

# PATTERN: CSV and Excel Processing

import csv
import os

# Data to write
sales_data = [
    {"product": "Laptop", "quantity": 5, "price": 1200},
    {"product": "Keyboard", "quantity": 10, "price": 75},
    {"product": "Monitor", "quantity": 3, "price": 300},
]
fieldnames = ["product", "quantity", "price"]

# Pattern: Write data using csv.DictWriter
with open("sales_report.csv", "w", newline="") as csvfile:
    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(sales_data)

os.remove("sales_report.csv") # Clean up

# PATTERN: CSV and Excel Processing

import openpyxl
import os

# Setup: Create a dummy Excel file
wb = openpyxl.Workbook()
ws = wb.active
ws.append(['Name', 'Age'])
ws.append(['Alice', 30])
wb.save("people_data.xlsx")

# Pattern: Load Workbook and read data
loaded_wb = openpyxl.load_workbook("people_data.xlsx")
sheet = loaded_wb.active
data_row = [cell.value for cell in sheet[2]] # Read second row

os.remove("people_data.xlsx") # Clean up

# PATTERN: CSV and Excel Processing

import openpyxl
import os

# Pattern: Create a new workbook and write data
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Inventory"

# Write headers
sheet.append(["Item", "Stock"])

# Write data rows
sheet.append(["Apples", 150])
sheet.append(["Bananas", 200])

workbook.save("stock_report.xlsx")

os.remove("stock_report.xlsx") # Clean up